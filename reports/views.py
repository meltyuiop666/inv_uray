from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from datetime import date
from django.http import HttpResponse
from transactions.models import BarangMasuk, BarangKeluar

# Export Libraries
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

@login_required
def laporan_masuk(request):
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Default to current month start and today if dates are not provided
    if not start_date:
        start_date = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = date.today().strftime('%Y-%m-%d')
        
    items = BarangMasuk.objects.filter(tanggal__range=[start_date, end_date]).select_related('barang', 'supplier').order_by('tanggal', 'created_at')
    total_qty = items.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
    
    context = {
        'items': items,
        'start_date': start_date,
        'end_date': end_date,
        'total_qty': total_qty,
        'print_date': date.today(),
    }
    return render(request, 'reports/masuk.html', context)

@login_required
def laporan_keluar(request):
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Default to current month start and today if dates are not provided
    if not start_date:
        start_date = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = date.today().strftime('%Y-%m-%d')
        
    items = BarangKeluar.objects.filter(tanggal__range=[start_date, end_date]).select_related('barang').order_by('tanggal', 'created_at')
    total_qty = items.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
    
    context = {
        'items': items,
        'start_date': start_date,
        'end_date': end_date,
        'total_qty': total_qty,
        'print_date': date.today(),
    }
    return render(request, 'reports/keluar.html', context)


# --- EXPORT EXCEL ---

@login_required
def export_masuk_excel(request):
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    if not start_date:
        start_date = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = date.today().strftime('%Y-%m-%d')
        
    items = BarangMasuk.objects.filter(tanggal__range=[start_date, end_date]).select_related('barang', 'supplier').order_by('tanggal', 'created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Barang Masuk"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws['A1'] = "LAPORAN BARANG MASUK"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="E11D48") # Rose-600
    
    ws['A2'] = f"Lookatstore.id Mempawah | Periode: {start_date} s.d {end_date}"
    ws['A2'].font = Font(name='Arial', size=11, italic=True)
    
    # Headers
    headers = ['No', 'Tanggal', 'Kode Barang', 'Nama Produk', 'Supplier', 'Jumlah', 'Keterangan']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid") # Rose-600
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Write Data
    row_num = 5
    for idx, item in enumerate(items, 1):
        ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=item.tanggal.strftime('%d/%m/%Y')).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=3, value=item.barang.kode_barang).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=item.barang.nama_produk)
        ws.cell(row=row_num, column=5, value=item.supplier.nama if item.supplier else "-")
        ws.cell(row=row_num, column=6, value=item.jumlah).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=7, value=item.keterangan or "-")
        
        # Apply style to data rows
        for col_idx in range(1, 8):
            c = ws.cell(row=row_num, column=col_idx)
            c.font = Font(name='Arial', size=10)
            c.border = thin_border
            if row_num % 2 == 0:
                c.fill = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid") # brand-50 alternating
        row_num += 1
        
    # Total row
    total_cell_label = ws.cell(row=row_num, column=5, value="Total")
    total_cell_label.font = Font(name='Arial', size=11, bold=True)
    total_cell_label.alignment = Alignment(horizontal="right")
    
    total_val = items.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
    total_cell_val = ws.cell(row=row_num, column=6, value=total_val)
    total_cell_val.font = Font(name='Arial', size=11, bold=True)
    total_cell_val.alignment = Alignment(horizontal="center")
    
    # Border for total row
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    for col_idx in range(1, 8):
        ws.cell(row=row_num, column=col_idx).border = double_bottom_border

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=laporan_barang_masuk_{start_date}_to_{end_date}.xlsx'
    wb.save(response)
    return response

@login_required
def export_keluar_excel(request):
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    if not start_date:
        start_date = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = date.today().strftime('%Y-%m-%d')
        
    items = BarangKeluar.objects.filter(tanggal__range=[start_date, end_date]).select_related('barang').order_by('tanggal', 'created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Barang Keluar"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws['A1'] = "LAPORAN BARANG KELUAR"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="E11D48") # Rose-600
    
    ws['A2'] = f"Lookatstore.id Mempawah | Periode: {start_date} s.d {end_date}"
    ws['A2'].font = Font(name='Arial', size=11, italic=True)
    
    # Headers
    headers = ['No', 'Tanggal', 'Kode Barang', 'Nama Produk', 'Alasan Keluar', 'Jumlah']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid") # Rose-600
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Write Data
    row_num = 5
    for idx, item in enumerate(items, 1):
        ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=item.tanggal.strftime('%d/%m/%Y')).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=3, value=item.barang.kode_barang).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=item.barang.nama_produk)
        ws.cell(row=row_num, column=5, value=item.keterangan or "-")
        ws.cell(row=row_num, column=6, value=item.jumlah).alignment = Alignment(horizontal="center")
        
        # Apply style to data rows
        for col_idx in range(1, 7):
            c = ws.cell(row=row_num, column=col_idx)
            c.font = Font(name='Arial', size=10)
            c.border = thin_border
            if row_num % 2 == 0:
                c.fill = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid")
        row_num += 1
        
    # Total row
    total_cell_label = ws.cell(row=row_num, column=5, value="Total")
    total_cell_label.font = Font(name='Arial', size=11, bold=True)
    total_cell_label.alignment = Alignment(horizontal="right")
    
    total_val = items.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
    total_cell_val = ws.cell(row=row_num, column=6, value=total_val)
    total_cell_val.font = Font(name='Arial', size=11, bold=True)
    total_cell_val.alignment = Alignment(horizontal="center")
    
    # Border for total row
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    for col_idx in range(1, 7):
        ws.cell(row=row_num, column=col_idx).border = double_bottom_border

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=laporan_barang_keluar_{start_date}_to_{end_date}.xlsx'
    wb.save(response)
    return response


# --- EXPORT PDF ---

@login_required
def export_masuk_pdf(request):
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    if not start_date:
        start_date = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = date.today().strftime('%Y-%m-%d')
        
    items = BarangMasuk.objects.filter(tanggal__range=[start_date, end_date]).select_related('barang', 'supplier').order_by('tanggal', 'created_at')
    total_qty = items.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=laporan_barang_masuk_{start_date}_to_{end_date}.pdf'
    
    # Prepare PDF layout
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=40, bottomMargin=40)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#BE123C'), # brand-700 / Rose
        alignment=1 # Center
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#475569'), # Slate-600
        alignment=1 # Center
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12
    )
    cell_style_bold = ParagraphStyle(
        'CellTextBold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    # Header Info
    story.append(Paragraph("LAPORAN TRANSAKSI BARANG MASUK", title_style))
    story.append(Paragraph("Lookatstore.id Mempawah - Busana & Fashion Wanita", subtitle_style))
    story.append(Paragraph(f"Periode Tanggal: {start_date} s.d. {end_date}", subtitle_style))
    story.append(Spacer(1, 15))
    
    # Table Columns Widths (Total: A4 usable width is ~523 pt)
    # No (30), Tanggal (60), Kode (65), Nama Produk (138), Supplier (95), Jumlah (45), Keterangan (90)
    col_widths = [25, 60, 65, 138, 95, 45, 95]
    
    table_data = [[
        Paragraph('No', header_style),
        Paragraph('Tanggal', header_style),
        Paragraph('Kode', header_style),
        Paragraph('Nama Produk', header_style),
        Paragraph('Supplier', header_style),
        Paragraph('Jumlah', header_style),
        Paragraph('Keterangan', header_style)
    ]]
    
    for idx, item in enumerate(items, 1):
        table_data.append([
            Paragraph(str(idx), cell_style),
            Paragraph(item.tanggal.strftime('%d/%m/%Y'), cell_style),
            Paragraph(item.barang.kode_barang, cell_style),
            Paragraph(item.barang.nama_produk, cell_style),
            Paragraph(item.supplier.nama if item.supplier else '-', cell_style),
            Paragraph(str(item.jumlah), cell_style),
            Paragraph(item.keterangan or '-', cell_style),
        ])
        
    # Total row
    table_data.append([
        Paragraph('', cell_style),
        Paragraph('', cell_style),
        Paragraph('', cell_style),
        Paragraph('', cell_style),
        Paragraph('Total', cell_style_bold),
        Paragraph(str(total_qty), cell_style_bold),
        Paragraph('', cell_style)
    ])
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Styling Table
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E11D48')), # Rose-600
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
    ])
    
    # Alternating row colors
    for r in range(1, len(table_data) - 1):
        if r % 2 == 0:
            t_style.add('BACKGROUND', (0, r), (-1, r), colors.HexColor('#FFF1F2'))
            
    # Styling Total Row
    t_style.add('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#000000'))
    t_style.add('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#000000'))
    
    t.setStyle(t_style)
    story.append(t)
    
    # Signature Section
    story.append(Spacer(1, 35))
    sig_style = ParagraphStyle(
        'SigText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#1E293B')
    )
    
    sig_data = [
        [Paragraph("", sig_style), Paragraph(f"Mempawah, {date.today().strftime('%d %B %Y')}<br/>Pemilik Lookatstore.id<br/><br/><br/><br/><br/>( ________________________ )", sig_style)]
    ]
    sig_table = Table(sig_data, colWidths=[300, 223])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    
    story.append(sig_table)
    
    doc.build(story)
    return response

@login_required
def export_keluar_pdf(request):
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    if not start_date:
        start_date = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = date.today().strftime('%Y-%m-%d')
        
    items = BarangKeluar.objects.filter(tanggal__range=[start_date, end_date]).select_related('barang').order_by('tanggal', 'created_at')
    total_qty = items.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=laporan_barang_keluar_{start_date}_to_{end_date}.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#BE123C'),
        alignment=1
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#475569'),
        alignment=1
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12
    )
    cell_style_bold = ParagraphStyle(
        'CellTextBold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white,
        alignment=1
    )
    
    story.append(Paragraph("LAPORAN TRANSAKSI BARANG KELUAR", title_style))
    story.append(Paragraph("Lookatstore.id Mempawah - Busana & Fashion Wanita", subtitle_style))
    story.append(Paragraph(f"Periode Tanggal: {start_date} s.d. {end_date}", subtitle_style))
    story.append(Spacer(1, 15))
    
    # Table Columns Widths (Total: A4 usable width is ~523 pt)
    # No (30), Tanggal (75), Kode (80), Nama Produk (180), Alasan Keluar (100), Jumlah (58)
    col_widths = [30, 75, 80, 180, 100, 58]
    
    table_data = [[
        Paragraph('No', header_style),
        Paragraph('Tanggal', header_style),
        Paragraph('Kode', header_style),
        Paragraph('Nama Produk', header_style),
        Paragraph('Alasan Keluar', header_style),
        Paragraph('Jumlah', header_style)
    ]]
    
    for idx, item in enumerate(items, 1):
        table_data.append([
            Paragraph(str(idx), cell_style),
            Paragraph(item.tanggal.strftime('%d/%m/%Y'), cell_style),
            Paragraph(item.barang.kode_barang, cell_style),
            Paragraph(item.barang.nama_produk, cell_style),
            Paragraph(item.keterangan or '-', cell_style),
            Paragraph(str(item.jumlah), cell_style),
        ])
        
    # Total row
    table_data.append([
        Paragraph('', cell_style),
        Paragraph('', cell_style),
        Paragraph('', cell_style),
        Paragraph('', cell_style),
        Paragraph('Total', cell_style_bold),
        Paragraph(str(total_qty), cell_style_bold),
    ])
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E11D48')), # Rose-600
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
    ])
    
    for r in range(1, len(table_data) - 1):
        if r % 2 == 0:
            t_style.add('BACKGROUND', (0, r), (-1, r), colors.HexColor('#FFF1F2'))
            
    t_style.add('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#000000'))
    t_style.add('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#000000'))
    
    t.setStyle(t_style)
    story.append(t)
    
    story.append(Spacer(1, 35))
    sig_style = ParagraphStyle(
        'SigText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#1E293B')
    )
    
    sig_data = [
        [Paragraph("", sig_style), Paragraph(f"Mempawah, {date.today().strftime('%d %B %Y')}<br/>Pemilik Lookatstore.id<br/><br/><br/><br/><br/>( ________________________ )", sig_style)]
    ]
    sig_table = Table(sig_data, colWidths=[300, 223])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    
    story.append(sig_table)
    
    doc.build(story)
    return response
